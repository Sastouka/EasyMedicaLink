import os
import re
import uuid
import shutil
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional

import pandas as pd
from fpdf import FPDF
from openpyxl import Workbook
from flask import (
    Blueprint, request, render_template_string,
    redirect, url_for, session, jsonify, send_file
)
import utils
import theme
import login

# These variables will be dynamically defined once set_dynamic_base_dir is called
EXCEL_DIR: Optional[Path] = None
PDF_DIR: Optional[Path] = None
EXCEL_FILE: Optional[Path] = None
CONSULT_FILE: Optional[Path] = None
BASE_PATIENT_FILE: Optional[Path] = None

# ------------------------------------------------------------------
# DIRECTORY CONFIGURATION
# ------------------------------------------------------------------
def set_rdv_dirs():
    """Sets the dynamic directory paths for RDV blueprint."""
    global EXCEL_DIR, PDF_DIR, EXCEL_FILE, CONSULT_FILE, BASE_PATIENT_FILE

    # Ensure utils.EXCEL_FOLDER and utils.PDF_FOLDER are set by set_dynamic_base_dir
    if utils.EXCEL_FOLDER is None or utils.PDF_FOLDER is None:
        print("ERROR: utils.EXCEL_FOLDER or utils.PDF_FOLDER not set. Cannot initialize RDV paths.")
        return

    EXCEL_DIR = Path(utils.EXCEL_FOLDER)
    PDF_DIR   = Path(utils.PDF_FOLDER)
    os.makedirs(EXCEL_DIR, exist_ok=True)
    os.makedirs(PDF_DIR, exist_ok=True)

    EXCEL_FILE        = EXCEL_DIR / "DonneesRDV.xlsx"
    CONSULT_FILE      = EXCEL_DIR / "ConsultationData.xlsx"
    BASE_PATIENT_FILE = EXCEL_DIR / "info_Base_patient.xlsx"
    print(f"DEBUG: RDV paths set. EXCEL_DIR: {EXCEL_DIR}, PDF_DIR: {PDF_DIR}")

def backup_info_base_patient():
    # Ensure BASE_PATIENT_FILE and EXCEL_FOLDER are set
    if BASE_PATIENT_FILE is None or utils.EXCEL_FOLDER is None:
        print("ERROR: BASE_PATIENT_FILE or utils.EXCEL_FOLDER not set. Cannot backup patient file.")
        return

    source_file = str(BASE_PATIENT_FILE)
    if os.path.exists(source_file):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(
            utils.EXCEL_FOLDER,
            f"backup_info_Base_patient_{timestamp}.xlsx"
        )
        shutil.copy2(source_file, backup_file)

def initialize_base_patient_file():
    """Initialises the info_Base_patient.xlsx file with unified columns."""
    if BASE_PATIENT_FILE is None:
        print("ERROR: BASE_PATIENT_FILE not set. Cannot initialize base patient file.")
        return

    wb = Workbook()
    sheet = wb.active
    sheet.title = "BasePatients"
    sheet.append([
        "ID", "Nom", "Prenom", "DateNaissance", "Sexe", "Âge",
        "Antécédents", "Téléphone"
    ])
    wb.save(BASE_PATIENT_FILE)
    print(f"DEBUG: Fichier info_Base_patient.xlsx initialisé avec les colonnes unifiées.")

def save_base_patient_df(df_new: pd.DataFrame):
    """Saves or updates data in info_Base_patient.xlsx."""
    if BASE_PATIENT_FILE is None:
        print("ERROR: BASE_PATIENT_FILE not set. Cannot save base patient dataframe.")
        return

    if not BASE_PATIENT_FILE.exists():
        initialize_base_patient_file()
        df_existing = pd.DataFrame(columns=[
            "ID", "Nom", "Prenom", "DateNaissance", "Sexe", "Âge",
            "Antécédents", "Téléphone"
        ])
    else:
        df_existing = pd.read_excel(BASE_PATIENT_FILE, dtype=str).fillna('')

    expected_cols = ["ID", "Nom", "Prenom", "DateNaissance", "Sexe", "Âge", "Antécédents", "Téléphone"]
    for col in expected_cols:
        if col not in df_new.columns:
            df_new[col] = ''

    df_new_filtered = df_new[expected_cols]

    df_combined = pd.concat([df_existing, df_new_filtered], ignore_index=True)
    if "ID" in df_combined.columns:
        df_combined.drop_duplicates(subset=["ID"], keep="last", inplace=True)
    
    df_combined.to_excel(BASE_PATIENT_FILE, index=False)
    print(f"DEBUG: Données sauvegardées dans info_Base_patient.xlsx. Total d'entrées: {len(df_combined)}")

# ------------------------------------------------------------------
# EXCEL HELPER (for DonneesRDV.xlsx)
# ------------------------------------------------------------------
def initialize_excel_file():
    """Initialises the DonneesRDV.xlsx file with unified columns."""
    if EXCEL_FILE is None:
        print("ERROR: EXCEL_FILE not set. Cannot initialize excel file.")
        return

    wb = Workbook()
    sheet = wb.active
    sheet.title = "RDV"
    sheet.append([
        "Num Ordre", "ID", "Nom", "Prenom", "DateNaissance", "Sexe", "Âge",
        "Antécédents", "Téléphone", "Date", "Heure", "Medecin_Email"
    ])
    wb.save(EXCEL_FILE)
    print(f"DEBUG: Fichier DonneesRDV.xlsx initialisé avec les colonnes unifiées.")

def load_df() -> pd.DataFrame:
    """Loads the DataFrame from DonneesRDV.xlsx, adding missing columns if any."""
    if EXCEL_FILE is None:
        print("ERROR: EXCEL_FILE not set. Cannot load dataframe.")
        return pd.DataFrame()

    if not EXCEL_FILE.exists():
        initialize_excel_file()
    df = pd.read_excel(EXCEL_FILE, dtype=str).fillna('')
    if 'Nom' not in df.columns:
        df.insert(loc=2, column='Nom', value='')
    if 'Prenom' not in df.columns:
        df.insert(loc=3, column='Prenom', value='')
    if 'DateNaissance' not in df.columns:
        df.insert(loc=4, column='DateNaissance', value='')
    if 'Sexe' not in df.columns:
        df.insert(loc=5, column='Sexe', value='')
    if 'Âge' not in df.columns:
        df.insert(loc=6, column='Âge', value='')
    if 'Antécédents' not in df.columns:
        df.insert(loc=7, column='Antécédents', value='')
    if 'Téléphone' not in df.columns:
        df.insert(loc=8, column='Téléphone', value='')
    if 'Medecin_Email' not in df.columns:
        df['Medecin_Email'] = ''
    return df

def save_df(df: pd.DataFrame):
    """Saves the DataFrame to DonneesRDV.xlsx."""
    if EXCEL_FILE is None:
        print("ERROR: EXCEL_FILE not set. Cannot save dataframe.")
        return
    df.to_excel(EXCEL_FILE, index=False)

def load_patients() -> dict:
    """Loads patients from DonneesRDV.xlsx for the datalist (patient_id)."""
    patients = {}
    df = load_df()
    for _, row in df.iterrows():
        pid = str(row["ID"]).strip()
        if not pid:
            continue
        nom = str(row.get('Nom', '')).strip()
        prenom = str(row.get('Prenom', '')).strip()
        full_name = f"{nom} {prenom}".strip()
        patients[pid] = {
            "name":          full_name,
            "nom":           nom,
            "prenom":        prenom,
            "date_of_birth": str(row.get("DateNaissance", "")),
            "gender":        str(row.get("Sexe", "")),
            "age":           str(row.get("Âge", "")),
            "phone":         str(row.get("Téléphone", "")),
            "antecedents":   str(row.get("Antécédents", "")),
        }
    return patients

# ------------------------------------------------------------------
# CALCULATIONS / VALIDATIONS
# ------------------------------------------------------------------
def compute_age_str(dob: date) -> str:
    today_date = date.today()
    years = today_date.year - dob.year - (
        (today_date.month, today_date.day) < (dob.month, dob.day)
    )
    months = today_date.month - dob.month - (today_date.day < dob.day)
    if months < 0:
        months += 12
    return f"{years} ans {months} mois"

PHONE_RE = re.compile(r"^[+0]\d{6,14}$")

# ------------------------------------------------------------------
# BLUEPRINT DECLARATION
# ------------------------------------------------------------------
rdv_bp = Blueprint("rdv", __name__, url_prefix="/rdv")

# ------------------------------------------------------------------
# TRANSFER TO CONSULTATION ROUTE
# ------------------------------------------------------------------
@rdv_bp.route("/consult/<int:index>", methods=["GET", "POST"])
def consult_rdv(index):
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    df_rdv = load_df()
    if not (0 <= index < len(df_rdv)):
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'warning',
              title: 'Sélection invalide',
              text: 'Le rendez-vous que vous tentez de consulter n\'existe pas.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)
    rdv_row = df_rdv.iloc[index]

    cols = [
        "consultation_date", "patient_id", "patient_name", "nom", "prenom",
        "date_of_birth", "gender", "age", "patient_phone", "antecedents",
        "clinical_signs", "bp", "temperature", "heart_rate",
        "respiratory_rate", "diagnosis", "medications", "analyses", "radiologies",
        "certificate_category", "certificate_content", "rest_duration", "doctor_comment",
        "consultation_id", "Medecin_Email"
    ]

    if CONSULT_FILE is None:
        print("ERROR: CONSULT_FILE not set. Cannot access ConsultationData.xlsx.")
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'error',
              title: 'Erreur',
              text: 'Impossible d\'accéder aux données de consultation.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)

    if CONSULT_FILE.exists():
        df_consult = pd.read_excel(CONSULT_FILE, dtype=str).fillna('')
        for col in ["nom", "prenom", "Medecin_Email"]:
            if col not in df_consult.columns:
                df_consult[col] = ''
    else:
        df_consult = pd.DataFrame(columns=cols)

    form      = request.form
    med_list  = form.getlist("medications_list")
    anal_list = form.getlist("analyses_list")
    rad_list = form.getlist("radiologies_list")

    consultation_date = datetime.now().strftime("%Y-%m-%d")

    nom_from_rdv = rdv_row.get("Nom", "").strip()
    prenom_from_rdv = rdv_row.get("Prenom", "").strip()
    patient_full_name_from_rdv = f"{nom_from_rdv} {prenom_from_rdv}".strip()
    medecin_email_from_rdv = rdv_row.get("Medecin_Email", "").strip()

    new_row = {
        "consultation_date":    consultation_date,
        "patient_id":           rdv_row["ID"],
        "patient_name":         patient_full_name_from_rdv,
        "nom":                  nom_from_rdv,
        "prenom":               prenom_from_rdv,
        "date_of_birth":        form.get("patient_dob", rdv_row.get("DateNaissance", "")).strip(),
        "gender":               form.get("gender", rdv_row.get("Sexe", "")).strip(),
        "age":                  form.get("patient_age", rdv_row.get("Âge", "")).strip(),
        "patient_phone":        form.get("patient_phone", rdv_row.get("Téléphone", "")).strip(),
        "antecedents":          form.get("antecedents", rdv_row.get("Antécédents", "")).strip(),
        "clinical_signs":       form.get("clinical_signs", "").strip(),
        "bp":                   form.get("bp", "").strip(),
        "temperature":          form.get("temperature", "").strip(),
        "heart_rate":           form.get("heart_rate", "").strip(),
        "respiratory_rate":     form.get("respiratory_rate", "").strip(),
        "diagnosis":            form.get("diagnosis", "").strip(),
        "medications":          "; ".join(med_list),
        "analyses":             "; ".join(anal_list),
        "radiologies":          "; ".join(rad_list),
        "certificate_category": form.get("certificate_category", "").strip(),
        "certificate_content":  form.get("certificate_content", "").strip(),
        "rest_duration":        utils.extract_rest_duration(form.get("certificate_content", "")),
        "doctor_comment":       form.get("doctor_comment", "").strip(),
        "consultation_id":      str(uuid.uuid4()),
        "Medecin_Email":        medecin_email_from_rdv
    }

    for col in new_row.keys():
        if col not in df_consult.columns:
            df_consult[col] = ''

    df_consult = pd.concat([df_consult, pd.DataFrame([new_row])], ignore_index=True)
    df_consult.to_excel(CONSULT_FILE, index=False)

    patient_base_data_from_rdv = pd.DataFrame([{
        "ID": rdv_row["ID"],
        "Nom": rdv_row["Nom"],
        "Prenom": rdv_row["Prenom"],
        "DateNaissance": rdv_row["DateNaissance"],
        "Sexe": rdv_row["Sexe"],
        "Âge": rdv_row["Âge"],
        "Antécédents": rdv_row["Antécédents"],
        "Téléphone": rdv_row["Téléphone"]
    }])
    save_base_patient_df(patient_base_data_from_rdv)

    df_rdv = df_rdv.drop(df_rdv.index[index]).reset_index(drop=True)
    save_df(df_rdv)

    return render_template_string("""
    <!DOCTYPE html><html><head>
      <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
    </head><body>
      <script>
        Swal.fire({
          icon: 'success',
          title: 'Consultation enregistrée',
          text: 'Patient transféré et consultation enregistrée ✔',
          confirmButtonText: 'OK'
        }).then(() => {
          window.location.href = '{{ url_for("rdv.rdv_home") }}';
        });
      </script>
    </body></html>
    """)

# ------------------------------------------------------------------
# MAIN RDV ROUTE
# ------------------------------------------------------------------
@rdv_bp.route("/", methods=["GET", "POST"])
def rdv_home():
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    config       = utils.load_config()
    session['theme'] = config.get('theme', theme.DEFAULT_THEME)
    theme_vars   = theme.current_theme()
    theme_names  = list(theme.THEMES.keys())

    df       = load_df()
    patients = load_base_patients()

    # Charger les médecins gérés par l'administrateur courant
    doctors = []
    config_doctor_name = config.get('doctor_name', '').strip()
    
    # 1. Ajouter le médecin configuré dans la config de l'administrateur
    if config_doctor_name:
        name_parts = config_doctor_name.split(' ')
        prenom_config = ' '.join(name_parts[:-1]) if len(name_parts) > 1 else ''
        nom_config = name_parts[-1] if name_parts else ''
        
        # Utilisez l'email de l'admin comme email du médecin configuré
        doctors.append({'email': admin_email_from_session, 'nom': nom_config, 'prenom': prenom_config})

    # 2. Ajouter les médecins ayant des comptes
    all_users = login.load_users()
    for email, user_data in all_users.items():
        # Vérifiez que l'utilisateur est un médecin et appartient à l'administrateur actuel
        if user_data.get('role') == 'medecin' and user_data.get('owner') == admin_email_from_session:
            # Éviter les doublons si l'e-mail du médecin correspond à l'e-mail de l'admin déjà ajouté
            is_duplicate = False
            for doc in doctors:
                if doc['email'] == email:
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                doctors.append({'email': email, 'nom': user_data.get('nom'), 'prenom': user_data.get('prenom')})

    # Trier la liste des médecins par prénom puis nom pour un affichage cohérent
    doctors = sorted(doctors, key=lambda x: (x.get('prenom', ''), x.get('nom', '')))

    rdv_start_time = config.get('rdv_start_time', '08:00')
    rdv_end_time = config.get('rdv_end_time', '17:45')
    rdv_interval_minutes = config.get('rdv_interval_minutes', 15)

    all_available_timeslots = utils.generate_time_slots(
        rdv_start_time, rdv_end_time, rdv_interval_minutes
    )

    filt_date = request.args.get("date", "")
    iso_today = datetime.now().strftime("%Y-%m-%d")

    current_date_for_slots = filt_date if filt_date else iso_today

    reserved_slots = df[df["Date"] == current_date_for_slots]["Heure"].tolist()
    print(f"DEBUG: Créneaux réservés pour {current_date_for_slots} : {reserved_slots}")

    if request.method == "POST":
        f         = request.form
        pid       = f.get("patient_id", "").strip()
        nom       = f.get("patient_nom", "").strip()
        prenom    = f.get("patient_prenom", "").strip()
        gender    = f.get("patient_gender", "").strip()
        dob_str   = f.get("patient_dob", "").strip()
        ant       = f.get("patient_ant", "").strip()
        phone     = f.get("patient_phone", "").strip()
        date_rdv  = f.get("rdv_date", "").strip()
        time_rdv  = f.get("rdv_time", "").strip()
        medecin_email = f.get("medecin_select", "").strip()

        if ((df["Date"] == date_rdv) & (df["Heure"] == time_rdv) & (df["Medecin_Email"] == medecin_email)).any():
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Créneau déjà réservé',
                  text: 'Le rendez-vous du {{ date_rdv }} à {{ time_rdv }} pour le médecin sélectionné existe déjà.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """, date_rdv=date_rdv, time_rdv=time_rdv)

        # 2. NOUVELLE VÉRIFICATION (MISE À JOUR)
        # Un patient ne peut pas avoir 2 RDV le même jour AVEC LE MÊME MÉDECIN
        if ((df["ID"] == pid) & (df["Date"] == date_rdv) & (df["Medecin_Email"] == medecin_email)).any():
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'error',
                  title: 'Rendez-vous Dupliqué',
                  text: 'Ce patient (ID: {{ pid }}) a déjà un rendez-vous programmé pour le {{ date_rdv }} avec ce médecin.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """, pid=pid, date_rdv=date_rdv)

        # 3. VÉRIFICATION DES CHAMPS MANQUANTS (inchangé)
        if not all([pid, nom, prenom, gender, dob_str, ant, phone, date_rdv, time_rdv, medecin_email]):
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Champs manquants',
                  text: 'Veuillez remplir tous les champs, y compris le médecin.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """)

        if not all([pid, nom, prenom, gender, dob_str, ant, phone, date_rdv, time_rdv, medecin_email]):
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Champs manquants',
                  text: 'Veuillez remplir tous les champs, y compris le médecin.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """)
        if not PHONE_RE.fullmatch(phone):
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Téléphone invalide',
                  text: 'Le format du numéro de téléphone est invalide.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """)

        try:
            dob_date = datetime.strptime(dob_str, "%Y-%m-%d").date()
            if dob_date > date.today():
                raise ValueError
        except ValueError:
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Date de naissance invalide',
                  text: 'La date de naissance est invalide ou dans le futur.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """)

        age_text = compute_age_str(dob_date)
        full_name = f"{nom} {prenom}".strip()
        
        if pid in patients and (
            patients[pid]["name"].lower() != full_name.lower()
        ):
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'error',
                  title: 'ID Patient existant',
                  text: 'L\'ID {{ pid }} appartient déjà à {{ patient_name_exist }}. Veuillez utiliser un autre ID ou corriger les informations du patient existant.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """, pid=pid, patient_name_exist=patients[pid]['name'])

        if pid in df["ID"].values:
            idx_to_update = df[df["ID"] == pid].index
            df.loc[idx_to_update, [
                "Nom", "Prenom", "DateNaissance", "Sexe", "Âge", "Téléphone", "Antécédents"
            ]] = [nom, prenom, dob_str, gender, age_text, phone, ant]
        
        num_ord = utils.calculate_order_number(
            time_rdv, rdv_start_time, rdv_interval_minutes
        )
        new_rdv_row_data = {
            "Num Ordre": num_ord,
            "ID": pid,
            "Nom": nom,
            "Prenom": prenom,
            "DateNaissance": dob_str,
            "Sexe": gender,
            "Âge": age_text,
            "Antécédents": ant,
            "Téléphone": phone,
            "Date": date_rdv,
            "Heure": time_rdv,
            "Medecin_Email": medecin_email
        }
        new_rdv_row_df = pd.DataFrame([new_rdv_row_data], columns=df.columns)

        df = pd.concat([df, new_rdv_row_df], ignore_index=True)
        df.drop_duplicates(subset=["ID","Date","Heure", "Medecin_Email"], keep="last", inplace=True)
        save_df(df)

        patient_base_data = pd.DataFrame([{
            "ID": pid,
            "Nom": nom,
            "Prenom": prenom,
            "DateNaissance": dob_str,
            "Sexe": gender,
            "Âge": age_text,
            "Antécédents": ant,
            "Téléphone": phone
        }])
        save_base_patient_df(patient_base_data)

        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'success',
              title: 'RDV confirmé ✅',
              text: 'Le rendez-vous a été enregistré avec succès.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)

    df_view   = df[df["Date"] == filt_date] if filt_date else df
    df_view   = df_view.sort_values("Num Ordre", key=lambda s: pd.to_numeric(s, errors="coerce"))

    # MODIFICATION COMMENCE ICI pour rdv_home: Préparez les données pour le tableau principal
    df_for_template = []
    for original_df_index, row_series in df_view.iterrows():
        row_dict = row_series.to_dict()
        row_dict['original_df_index'] = original_df_index # Ajoutez l'index original
        df_for_template.append(row_dict)
    # MODIFICATION FINIT ICI pour rdv_home

    # MODIFICATION COMMENCE ICI POUR today_rdv dans rdv_home
    temp_today_rdv_df = df[(df["Date"] == iso_today)].sort_values("Num Ordre", key=lambda s: pd.to_numeric(s, errors="coerce"))
    today_rdv_for_template = []
    for original_df_index_today, row_series_today in temp_today_rdv_df.iterrows():
        row_dict_today = row_series_today.to_dict()
        row_dict_today['original_df_index'] = original_df_index_today # Ajoutez l'index original
        today_rdv_for_template.append(row_dict_today)
    # MODIFICATION FINIT ICI POUR today_rdv dans rdv_home

    today     = datetime.now().strftime("%d/%m/%Y")

    logged_in_full_name = None
    user_email = session.get('email')

    if user_email:
        all_users_data = login.load_users()
        user_info = all_users_data.get(user_email)
        if user_info:
            logged_in_full_name = f"{user_info.get('prenom', '')} {user_info.get('nom', '')}".strip()
            if not logged_in_full_name:
                logged_in_full_name = None

    return render_template_string(
        rdv_template,
        config=config,
        theme_vars=theme_vars,
        theme_names=theme_names,
        patients=patients,
        doctors=doctors,
        df=df_for_template, # Passez la nouvelle liste de dictionnaires avec l'index original
        timeslots=all_available_timeslots,
        today=today,
        iso_today=iso_today,
        filt_date=filt_date,
        enumerate=enumerate,
        reserved_slots=reserved_slots,
        today_rdv=today_rdv_for_template, # Passez la nouvelle liste de dictionnaires pour le RDV du jour
        logged_in_doctor_name=logged_in_full_name
    )

# ------------------------------------------------------------------
# PRINT TODAY'S APPOINTMENTS
# ------------------------------------------------------------------
@rdv_bp.route("/pdf_today")
def pdf_today():
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    try:
        df = load_df()

        df["Date_str"] = df["Date"].astype(str)
        df["Date_parsed"] = pd.to_datetime(df["Date_str"], errors="coerce")
        
        df_valid_dates = df.dropna(subset=["Date_parsed"])
        df_valid_dates["Date_only"] = df_valid_dates["Date_parsed"].dt.date

        today = date.today()

        df_today = df_valid_dates[df_valid_dates["Date_only"] == today].copy()

        if df_today.empty:
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Aucun RDV',
                  text: 'Aucun rendez-vous n\'est prévu pour aujourd\'hui.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """)

        df_today["Num Ordre_numeric"] = pd.to_numeric(df_today["Num Ordre"], errors="coerce")
        df_today = df_today.sort_values(by=["Num Ordre_numeric", "Heure"]).drop(columns=["Num Ordre_numeric"])

        headers = ["ID", "Nom", "Prénom", "Âge", "Téléphone", "Antécédents", "Date", "Heure", "Médecin", "Num Ordre"]
        data = []
        admin_email = session.get('admin_email', 'default_admin@example.com')
        all_users = login.load_users().items()
        
        doctor_names_map = {
            u_email: f"{u_data.get('prenom', '').strip()} {u_data.get('nom', '').strip()}"
            for u_email, u_data in all_users if u_data.get('role') == 'medecin' and u_data.get('owner') == admin_email
        }
        # Inclure le médecin configuré dans le mapping s'il n'est pas déjà un utilisateur médecin
        config_doctor_name_from_config = config.get('doctor_name', '').strip()
        if config_doctor_name_from_config and admin_email not in doctor_names_map:
            doctor_names_map[admin_email] = config_doctor_name_from_config


        for _, row in df_today.iterrows():
            medecin_email = row.get("Medecin_Email", "")
            medecin_display_name = doctor_names_map.get(medecin_email, medecin_email if medecin_email else "N/A")
            data.append([
                row.get("ID", ""),
                row.get("Nom", ""),
                row.get("Prenom", ""),
                row.get("Âge", ""),
                row.get("Téléphone", ""),
                row.get("Antécédents", ""),
                row.get("Date", ""),
                row.get("Heure", ""),
                medecin_display_name,
                row.get("Num Ordre", "")
            ])

        today_str = today.strftime("%d-%m-%Y")
        
        if PDF_DIR is None:
            print("ERROR: PDF_DIR not set. Cannot save PDF.")
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'error',
                  title: 'Erreur PDF',
                  text: 'Impossible d\'accéder au répertoire PDF.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.rdv_home") }}';
                });
              </script>
            </body></html>
            """)

        pdf_path = PDF_DIR / f"RDV_du_{today_str.replace('-', '')}.pdf"
        
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        
        try:
            font_path = 'C:\\Windows\\Fonts\\arial.ttf'
            if not os.path.exists(font_path):
                print(f"WARNING: Arial font file not found at {font_path}. Attempting to use default FPDF fonts.")
                pdf.set_font("Helvetica", size=20)
            else:
                pdf.add_font('Arial', '', font_path, uni=True)
                pdf.set_font("Arial", size=20)
        except Exception as font_e:
            print(f"WARNING: Could not load Arial font from specified path. Falling back to default. Error: {font_e}")
            pdf.set_font("Helvetica", size=20)

        pdf.add_page()
        pdf.set_font("Arial", size=20, style='B')
        pdf.cell(0, 10, txt=f"RDV du {today.strftime('%d/%m/%Y')}", ln=True, align='C')
        pdf.ln(5)

        col_widths = calculate_pdf_column_widths(headers, data, pdf)

        pdf.set_font("Arial", size=12, style='B')
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, border=1, align='C')
        pdf.ln()

        pdf.set_font("Arial", size=12)
        for row in data:
            for i, item in enumerate(row):
                text = str(item)
                align = 'C' if headers[i] in ["ID", "Âge", "Téléphone", "Date", "Heure", "Num Ordre"] else 'L'
                pdf.cell(col_widths[i], 8, text, border=1, align=align)
            pdf.ln()

        pdf.output(str(pdf_path), 'F')
        return send_file(str(pdf_path), as_attachment=True, download_name=pdf_path.name)

    except Exception as e:
        print(f"ERROR: Full PDF generation error: {e}")
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'error',
              title: 'Erreur PDF',
              text: 'Erreur lors de la génération du PDF: {{ error_message }}',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """, error_message=str(e))

def calculate_pdf_column_widths(headers, data, pdf: FPDF):
    col_widths = [pdf.get_string_width(h) + 8 for h in headers]
    for row in data:
        for i, item in enumerate(row):
            w = pdf.get_string_width(str(item)) + 8
            if w > col_widths[i]:
                col_widths[i] = w
    page_width = pdf.w - pdf.l_margin - pdf.r_margin
    total = sum(col_widths)
    if total > page_width:
        scale = page_width / total
        col_widths = [w * scale for w in col_widths]
    
    min_col_width = page_width / len(headers) * 0.5
    col_widths = [max(w, min_col_width) for w in col_widths]

    total_after_min = sum(col_widths)
    if total_after_min > page_width:
        scale_after_min = page_width / total_after_min
        col_widths = [w * scale_after_min for w in col_widths]

    return col_widths

# ------------------------------------------------------------------
# MODIFY AN APPOINTMENT
# ------------------------------------------------------------------
@rdv_bp.route("/edit/<int:index>", methods=["GET", "POST"])
def edit_rdv(index):
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    config       = utils.load_config()
    session['theme'] = config.get('theme', theme.DEFAULT_THEME)
    theme_vars   = theme.current_theme()
    theme_names  = list(theme.THEMES.keys())
    patients     = load_patients()
    
    # Charger les médecins gérés par l'administrateur courant
    doctors = []
    config_doctor_name = config.get('doctor_name', '').strip()
    
    # 1. Ajouter le médecin configuré dans la config de l'administrateur
    if config_doctor_name:
        name_parts = config_doctor_name.split(' ')
        prenom_config = ' '.join(name_parts[:-1]) if len(name_parts) > 1 else ''
        nom_config = name_parts[-1] if name_parts else ''
        
        doctors.append({'email': admin_email_from_session, 'nom': nom_config, 'prenom': prenom_config})

    # 2. Ajouter les médecins ayant des comptes
    all_users = login.load_users()
    for email, user_data in all_users.items():
        if user_data.get('role') == 'medecin' and user_data.get('owner') == admin_email_from_session:
            is_duplicate = False
            for doc in doctors:
                if doc['email'] == email:
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                doctors.append({'email': email, 'nom': user_data.get('nom'), 'prenom': user_data.get('prenom')})

    # Trier la liste des médecins par prénom puis nom pour un affichage cohérent
    doctors = sorted(doctors, key=lambda x: (x.get('prenom', ''), x.get('nom', '')))

    rdv_start_time = config.get('rdv_start_time', '08:00')
    rdv_end_time = config.get('rdv_end_time', '17:45')
    rdv_interval_minutes = config.get('rdv_interval_minutes', 15)

    all_available_timeslots = utils.generate_time_slots(
        rdv_start_time, rdv_end_time, rdv_interval_minutes
    )

    df = load_df()
    filt_date = request.args.get("date", "")
    df_view   = df[df["Date"] == filt_date] if filt_date else df
    df_view   = df_view.sort_values("Num Ordre", key=lambda s: pd.to_numeric(s, errors="coerce"))
    today     = datetime.now().strftime("%d/%m/%Y")
    iso_today = datetime.now().strftime("%Y-%m-%d")

    if not (0 <= index < len(df)):
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'warning',
              title: 'RDV introuvable',
              text: 'Le rendez-vous que vous tentez de modifier n\'existe pas.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)

    if request.method == "POST":
        f = request.form
        required = [
            "patient_id","patient_nom","patient_prenom","patient_gender",
            "patient_dob","patient_ant","patient_phone","rdv_date","rdv_time",
            "medecin_select"
        ]
        if not all(f.get(k) for k in required):
            return render_template_string("""
            <!DOCTYPE html><html><head>
              <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
            </head><body>
              <script>
                Swal.fire({
                  icon: 'warning',
                  title: 'Champs manquants',
                  text: 'Veuillez remplir tous les champs du formulaire, y compris le médecin.',
                  confirmButtonText: 'OK'
                }).then(() => {
                  window.location.href = '{{ url_for("rdv.edit_rdv", index=index) }}';
                });
              </script>
            </body></html>
            """, index=index)

        dob_date = datetime.strptime(f["patient_dob"], "%Y-%m-%d").date()
        age_text = compute_age_str(dob_date)
        num_ord  = utils.calculate_order_number(
            f["rdv_time"], rdv_start_time, rdv_interval_minutes
        )

        df.at[index, "ID"]            = f["patient_id"]
        df.at[index, "Nom"]           = f["patient_nom"]
        df.at[index, "Prenom"]        = f["patient_prenom"]
        df.at[index, "Sexe"]          = f["patient_gender"]
        df.at[index, "DateNaissance"] = f["patient_dob"]
        df.at[index, "Âge"]           = age_text
        df.at[index, "Antécédents"]   = f["patient_ant"]
        df.at[index, "Téléphone"]     = f["patient_phone"]
        df.at[index, "Date"]          = f["rdv_date"]
        df.at[index, "Heure"]         = f["rdv_time"]
        df.at[index, "Num Ordre"]     = num_ord
        df.at[index, "Medecin_Email"] = f["medecin_select"]

        save_df(df)
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'success',
              title: 'RDV mis à jour ✅',
              text: 'Le rendez-vous a été mis à jour avec succès.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)

    edit_row = df.iloc[index]
    edit_date = edit_row["Date"]
    all_rdv = df.to_dict(orient="records")
    reserved_slots = [
        r["Heure"] for r in all_rdv
        if r["Date"] == edit_date and r["Heure"] != edit_row["Heure"] and r["Medecin_Email"] == edit_row["Medecin_Email"]
    ]

    # MODIFICATION COMMENCE ICI pour edit_rdv: Préparez les données pour le tableau principal
    df_for_edit_template = []
    for original_df_index, row_series in df_view.iterrows():
        row_dict = row_series.to_dict()
        row_dict['original_df_index'] = original_df_index # Ajoutez l'index original
        df_for_edit_template.append(row_dict)
    # MODIFICATION FINIT ICI pour edit_rdv

    # MODIFICATION COMMENCE ICI POUR today_rdv dans edit_rdv
    temp_today_rdv_df_edit = df[(df["Date"] == iso_today)].sort_values("Num Ordre", key=lambda s: pd.to_numeric(s, errors="coerce"))
    today_rdv_for_edit_template = []
    for original_df_index_today_edit, row_series_today_edit in temp_today_rdv_df_edit.iterrows():
        row_dict_today_edit = row_series_today_edit.to_dict()
        row_dict_today_edit['original_df_index'] = original_df_index_today_edit # Ajoutez l'index original
        today_rdv_for_edit_template.append(row_dict_today_edit)
    # MODIFICATION FINIT ICI POUR today_rdv dans edit_rdv

    return render_template_string(
        rdv_template,
        config=config,
        theme_vars=theme_vars,
        theme_names=theme_names,
        patients=patients,
        doctors=doctors,
        df=df_for_edit_template, # Utilisez la nouvelle liste de dictionnaires avec l'index original
        timeslots=all_available_timeslots,
        today=today,
        iso_today=iso_today,
        filt_date=filt_date,
        enumerate=enumerate,
        edit_index=index,
        edit_row=edit_row,
        reserved_slots=reserved_slots,
        today_rdv=today_rdv_for_edit_template # Utilisez la version corrigée de today_rdv
    )

# ------------------------------------------------------------------
# LOAD PATIENT_INFO
# ------------------------------------------------------------------
def load_base_patient_df() -> pd.DataFrame:
    if BASE_PATIENT_FILE is None:
        print("ERROR: BASE_PATIENT_FILE not set. Cannot load base patient dataframe.")
        return pd.DataFrame()
    if not BASE_PATIENT_FILE.exists():
        initialize_base_patient_file()
    return pd.read_excel(BASE_PATIENT_FILE, dtype=str)

def load_base_patients() -> dict:
    patients = {}
    df = load_base_patient_df()
    for _, row in df.iterrows():
        pid = str(row["ID"]).strip()
        if not pid:
            continue
        patients[pid] = {
            "name":          f"{row['Nom']} {row['Prenom']}".strip(),
            "nom":           str(row["Nom"]).strip(),
            "prenom":        str(row["Prenom"]).strip(),
            "date_of_birth": str(row["DateNaissance"]),
            "gender":        str(row["Sexe"]),
            "age":           str(row["Âge"]),
            "antecedents":   str(row["Antécédents"]),
            "phone":         str(row["Téléphone"]),
        }
    return patients

@rdv_bp.route("/patient_info/<patient_id>")
def patient_info(patient_id):
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    patients = load_base_patients()
    p = patients.get(patient_id)
    if not p:
        return jsonify({}), 404
    return jsonify(p)

# ------------------------------------------------------------------
# NEW ROUTE: Get reserved time slots for a given date (for AJAX)
# ------------------------------------------------------------------
@rdv_bp.route("/get_reserved_slots")
def get_reserved_slots():
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    config = utils.load_config()
    rdv_start_time = config.get('rdv_start_time', '08:00')
    rdv_end_time = config.get('rdv_end_time', '17:45')
    rdv_interval_minutes = config.get('rdv_interval_minutes', 15)

    date_param = request.args.get("date")
    medecin_email_param = request.args.get("medecin_email")

    if not date_param or not medecin_email_param:
        return jsonify({"error": "Date or doctor email parameter is missing"}), 400

    df = load_df()
    reserved_slots = df[(df["Date"] == date_param) & (df["Medecin_Email"] == medecin_email_param)]["Heure"].tolist()
    
    all_possible_slots = utils.generate_time_slots(
        rdv_start_time, rdv_end_time, rdv_interval_minutes
    )

    return jsonify({
        "reserved_slots": reserved_slots,
        "all_possible_slots": all_possible_slots
    })
    
@rdv_bp.route("/delete/<int:index>", methods=["POST"]) # MODIFIÉ: Ajout de methods=["POST"]
def delete_rdv(index):
    # Assurez-vous que les répertoires dynamiques sont définis
    admin_email_from_session = session.get('admin_email', 'default_admin@example.com')
    utils.set_dynamic_base_dir(admin_email_from_session)
    set_rdv_dirs()

    df = load_df()

    if not (0 <= index < len(df)):
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'warning',
              title: 'RDV introuvable',
              text: 'Le rendez-vous que vous tentez de supprimer n\'existe pas.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)

    try:
        df = df.drop(df.index[index]).reset_index(drop=True)
        save_df(df)
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'success',
              title: 'RDV supprimé ✅',
              text: 'Le rendez-vous a été supprimé avec succès.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)
    except Exception as e:
        print(f"ERROR: Erreur lors de la suppression du RDV à l'index {index}: {e}")
        return render_template_string("""
        <!DOCTYPE html><html><head>
          <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
        </head><body>
          <script>
            Swal.fire({
              icon: 'error',
              title: 'Erreur',
              text: 'Une erreur est survenue lors de la suppression du rendez-vous.',
              confirmButtonText: 'OK'
            }).then(() => {
              window.location.href = '{{ url_for("rdv.rdv_home") }}';
            });
          </script>
        </body></html>
        """)    

# ------------------------------------------------------------------
# JINJA TEMPLATE (responsive interface + theme menu + cards)
# ------------------------------------------------------------------
rdv_template = r"""
<!DOCTYPE html>
<html lang="fr">
{{ pwa_head()|safe }}
<head>
<link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&family=Great+Vibes&display=swap" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.datatables.net/1.13.7/css/dataTables.bootstrap5.min.css" rel="stylesheet">
<link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css" rel="stylesheet">
<script src="https://cdn.tailwindcss.com"></script>
<script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
<title>RDV – {{ config.nom_clinique or 'EasyMedicaLink' }}</title>

<style>
  :root {
    {% for var, val in theme_vars.items() %}
    --{{ var }}: {{ val }};
    {% endfor %}
    --font-primary: 'Poppins', sans-serif;
    --font-secondary: 'Great Vibes', cursive;
    --gradient-main: linear-gradient(45deg, var(--primary-color) 0%, var(--secondary-color) 100%);
    --shadow-light: 0 5px 15px rgba(0, 0, 0, 0.1);
    --shadow-medium: 0 8px 25px rgba(0, 0, 0, 0.2);
    --border-radius-lg: 1rem;
    --border-radius-md: 0.75rem;
    --border-radius-sm: 0.5rem;
  }

  body {
    font-family: var(--font-primary);
    background: var(--bg-color);
    color: var(--text-color);
    padding-top: 56px;
    transition: background 0.3s ease, color 0.3s ease;
  }

  .navbar {
    background: var(--gradient-main) !important;
    box-shadow: var(--shadow-medium);
  }
  .navbar-brand {
    font-family: var(--font-secondary);
    font-size: 2rem !important;
    color: white !important;
    display: flex;
    align-items: center;
    transition: transform 0.3s ease;
  }
  .navbar-brand:hover {
    transform: scale(1.05);
  }
  .navbar-toggler {
    border: none;
    outline: none;
  }
  .navbar-toggler i {
    color: white;
    font-size: 1.5rem;
  }

  .offcanvas-header {
    background: var(--gradient-main) !important;
    color: white;
  }
  .offcanvas-body {
    background: var(--card-bg) !important;
    color: var(--text-color) !important;
  }
  .offcanvas-title {
    font-weight: 600;
  }

  .card {
    border-radius: var(--border-radius-lg);
    box-shadow: var(--shadow-light);
    background: var(--card-bg) !important;
    color: var(--text-color) !important;
    border: none;
    transition: transform 0.2s ease, box-shadow 0.2s ease;
  }
  /* Removed translateY on card hover to prevent vibrations */
  .card:hover {
    box-shadow: var(--shadow-medium);
  }

  .card-header {
    background: var(--primary-color) !important;
    color: var(--button-text) !important;
    border-top-left-radius: var(--border-radius-lg);
    border-top-right-radius: var(--border-radius-lg);
    padding: 1.5rem;
    position: relative;
    overflow: hidden;
  }
  .card-header::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background: rgba(255, 255, 255, 0.1);
    transform: skewY(-5deg);
    transform-origin: top left;
    z-index: 0;
  }
  .card-header h1, .card-header .header-item, .card-header p {
    position: relative;
    z-index: 1;
    font-size: 1.8rem !important;
    font-weight: 700;
  }
  .card-header i {
    font-size: 1.8rem !important;
    margin-right: 0.5rem;
  }
  .header-item {
    font-size: 1.2rem !important;
    font-weight: 400;
  }

  /* Wizard Steps */
  .step-circle {
    width: 3rem;
    height: 3rem;
    border-radius: 50%;
    background: var(--secondary-color);
    color: #fff;
    display: flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
    font-size: 1.2rem;
    transition: all 0.3s ease;
    box-shadow: var(--shadow-light);
  }
  .step-circle.active {
    background: var(--primary-color);
    transform: scale(1.1);
    box-shadow: var(--shadow-medium);
  }
  .step-line {
    flex: 1;
    height: 4px;
    background: var(--secondary-color);
    transition: background 0.3s ease;
  }
  .step-line.active {
    background: var(--primary-color);
  }

  /* Floating Labels */
  .floating-label {
    position: relative;
    margin-bottom: 1rem; /* Optimized spacing */
  }
  .floating-label input,
  .floating-label select {
    padding: 1rem 0.75rem 0.5rem;
    height: auto;
    border-radius: var(--border-radius-sm);
    border: 1px solid var(--secondary-color);
    background-color: var(--card-bg);
    color: var(--text-color);
    transition: border-color 0.2s ease, box-shadow 0.2s ease;
  }
  .floating-label input:focus,
  .floating-label select:focus {
    border-color: var(--primary-color);
    box-shadow: 0 0 0 0.25rem rgba(var(--primary-color-rgb), 0.25);
    background-color: var(--card-bg); /* Keep background consistent on focus */
    color: var(--text-color);
  }
  .floating-label label {
    position: absolute;
    top: 0.75rem;
    left: 0.75rem;
    font-size: 1rem;
    color: var(--text-color-light); /* Lighter text for label */
    transition: all 0.2s ease;
    pointer-events: none; /* Allow clicks to pass through to input */
  }
  .floating-label input:focus + label,
  .floating-label input:not(:placeholder-shown) + label,
  .floating-label select:focus + label,
  .floating-label select:not([value=""]) + label {
    top: 0.25rem;
    left: 0.75rem;
    font-size: 0.75rem;
    color: var(--primary-color);
    background-color: var(--card-bg); /* Match input background */
    padding: 0 0.25rem;
    transform: translateX(-0.25rem); /* Adjust position slightly */
  }
  /* Specific style for date input placeholder */
  .floating-label input[type="date"]:not([value=""])::-webkit-datetime-edit-text,
  .floating-label input[type="date"]:not([value=""])::-webkit-datetime-edit-month-field,
  .floating-label input[type="date"]:not([value=""])::-webkit-datetime-edit-day-field,
  .floating-label input[type="date"]:not([value=""])::-webkit-datetime-edit-year-field {
    color: var(--text-color); /* Ensure date input text is visible */
  }
  .floating-label input[type="date"]::-webkit-calendar-picker-indicator {
    filter: {% if session.theme == 'dark' %}invert(1){% else %}none{% endif %}; /* Adjust calendar icon for dark theme */
  }


  /* Gender Radio Buttons */
  .gender-btn {
    display: flex;
    gap: 0.5rem;
    flex-wrap: wrap;
    margin-bottom: 1rem; /* Optimized spacing */
  }
  .gender-btn input {
    display: none;
  }
  .gender-btn label {
    flex: 1 1 calc(33.333% - 0.5rem);
    border: 2px solid var(--secondary-color);
    border-radius: var(--border-radius-md);
    padding: 0.75rem 0;
    cursor: pointer;
    transition: all 0.2s ease;
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 0.4rem;
    text-align: center;
    user-select: none;
    background-color: var(--card-bg);
    color: var(--text-color);
    font-weight: 600;
  }
  .gender-btn label:hover {
    background-color: rgba(var(--secondary-color-rgb), 0.1);
  }
  /* Removed translateY on gender radio button checked to prevent vibrations */
  .gender-btn input:checked + label {
    background: var(--gradient-main);
    color: var(--button-text);
    box-shadow: var(--shadow-medium);
    border-color: var(--primary-color);
  }

  /* Buttons */
  .btn {
    border-radius: var(--border-radius-md);
    font-weight: 600;
    transition: all 0.3s ease;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    padding: 0.75rem 1.25rem;
  }
  .btn i {
    margin-right: 0.5rem;
  }
  .btn-primary {
    background: var(--gradient-main);
    border: none;
    color: var(--button-text);
    box-shadow: var(--shadow-light);
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-primary:hover {
    box-shadow: var(--shadow-medium);
    background: var(--gradient-main); /* Ensure gradient persists on hover */
    opacity: 0.9;
  }
  .btn-success {
    background-color: var(--success-color);
    border-color: var(--success-color);
    color: white;
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-success:hover {
    background-color: var(--success-color-dark);
    border-color: var(--success-color-dark);
    box-shadow: var(--shadow-medium);
  }
  .btn-warning {
    background-color: var(--warning-color);
    border-color: var(--warning-color);
    color: white;
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-warning:hover {
    background-color: var(--warning-color-dark);
    border-color: var(--warning-color-dark);
    box-shadow: var(--shadow-medium);
  }
  .btn-danger {
    background-color: var(--danger-color);
    border-color: var(--danger-color);
    color: white;
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-danger:hover {
    background-color: var(--danger-color-dark);
    border-color: var(--danger-color-dark);
    box-shadow: var(--shadow-medium);
  }
  .btn-info { /* WhatsApp button */
    background-color: #25D366; /* WhatsApp green */
    border-color: #25D366;
    color: white;
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-info:hover {
    background-color: #1DA851;
    border-color: #1DA851;
    box-shadow: var(--shadow-medium);
  }
  .btn-outline-secondary {
    border-color: var(--secondary-color);
    color: var(--text-color); /* Changed to text-color for better visibility */
    background-color: transparent; /* Ensure transparent background */
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-outline-secondary:hover {
    background-color: var(--secondary-color);
    color: white;
    box-shadow: var(--shadow-light);
  }
  .btn-secondary { /* Added explicit styling for btn-secondary */
    background-color: var(--secondary-color);
    border-color: var(--secondary-color);
    color: var(--button-text); /* Changed to button-text for better visibility */
  }
  /* Removed translateY on button hover to prevent vibrations */
  .btn-secondary:hover {
    background-color: var(--secondary-color-dark); /* Assuming a dark variant exists in theme.py */
    border-color: var(--secondary-color-dark);
    box-shadow: var(--shadow-medium);
  }
  .btn-sm {
    padding: 0.5rem 0.8rem;
    font-size: 0.875rem;
  }

  /* DataTables */
  #rdvTable_wrapper .dataTables_filter input,
  #rdvTable_wrapper .dataTables_length select {
    border-radius: var(--border-radius-sm);
    border: 1px solid var(--secondary-color);
    padding: 0.5rem 0.75rem;
    background-color: var(--card-bg);
    color: var(--text-color);
  }
  #rdvTable_wrapper .dataTables_filter input:focus,
  #rdvTable_wrapper .dataTables_length select:focus {
    border-color: var(--primary-color);
    box_shadow: 0 0 0 0.25rem rgba(var(--primary-color-rgb), 0.25);
  }
  /* Hide the dropdown arrow for DataTables length select */
  #rdvTable_wrapper .dataTables_length select {
    -webkit-appearance: none; /* Remove default arrow for Chrome/Safari */
    -moz-appearance: none;    /* Remove default arrow for Firefox */
    appearance: none;         /* Remove default arrow for other browsers */
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 16 16'%3E%3Cpath fill='none' stroke='%23333' stroke-linecap='round' stroke-linejoin='round' stroke-width='2' d='M2 5l6 6 6-6'/%3E%3Csvg%3E"); /* Custom arrow */
    background-repeat: no-repeat;
    background-position: right 0.75rem center;
    background-size: 0.65em auto;
    padding-right: 2rem; /* Make space for the custom arrow */
  }
  /* Invert arrow color for dark theme */
  body.dark-theme #rdvTable_wrapper .dataTables_length select {
    background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 16 16'%3E%3Cpath fill='none' stroke='%23fff' stroke-linecap='round' stroke-linejoin='round' stroke-width='2' d='M2 5l6 6 6-6'/%3E%3Csvg%3E");
  }


  #rdvTable_wrapper .dataTables_paginate .pagination .page-item .page-link {
    border-radius: var(--border-radius-sm);
    margin: 0 0.2rem;
    background-color: var(--card-bg);
    color: var(--text-color);
    border: 1px solid var(--secondary-color);
  }
  #rdvTable_wrapper .dataTables_paginate .pagination .page-item.active .page-link {
    background: var(--gradient-main);
    border-color: var(--primary-color);
    color: var(--button-text);
  }
  #rdvTable_wrapper .dataTables_paginate .pagination .page-item .page-link:hover {
    background-color: rgba(var(--primary-color-rgb), 0.1);
    color: var(--primary-color);
  }
  .table {
    --bs-table-bg: var(--card-bg);
    --bs-table-color: var(--text-color);
    --bs-table-striped-bg: var(--table-striped-bg);
    --bs-table-striped-color: var(--text-color);
    --bs-table-border-color: var(--border-color);
  }
  .table thead th {
    background-color: var(--primary-color);
    color: var(--button-text);
    border-color: var(--primary-color);
  }
  .table tbody tr {
    transition: background-color 0.2s ease;
  }
  .table tbody tr:hover {
    background-color: rgba(var(--primary-color-rgb), 0.05) !important;
  }

  /* Flash messages */
  .alert {
    border-radius: var(--border-radius-md);
    font-weight: 600;
    position: fixed;
    top: 70px;
    left: 50%;
    transform: translateX(-50%);
    z-index: 1060; /* Above offcanvas */
    width: 90%;
    max-width: 500px;
    box-shadow: var(--shadow-medium);
    animation: fadeInOut 5s forwards;
  }

  @keyframes fadeInOut {
    0% { opacity: 0; transform: translateX(-50%) translateY(-20px); }
    10% { opacity: 1; transform: translateX(-50%) translateY(0); }
    90% { opacity: 1; transform: translateX(-50%) translateY(0); }
    100% { opacity: 0; transform: translateX(-50%) translateY(-20px); }
  }

  /* Footer */
  footer {
    background: var(--gradient-main);
    color: white;
    font-weight: 300;
    box-shadow: 0 -5px 15px rgba(0, 0, 0, 0.1);
  }
  footer a {
    color: white;
    text-decoration: none;
    transition: color 0.2s ease;
  }
  footer a:hover {
    color: var(--text-color-light);
  }

  /* New Calendar Card Styling */
  .calendar-card {
    border-radius: var(--border-radius-lg);
    box-shadow: var(--shadow-light);
    background: var(--card-bg);
    padding: 1.5rem;
    margin-bottom: 1.5rem;
  }

  .calendar-slot {
    display: flex;
    justify-content: space-between;
    align-items: center;
    padding: 0.75rem 1rem;
    margin-bottom: 0.5rem;
    border-radius: var(--border-radius-sm);
    background-color: rgba(var(--primary-color-rgb), 0.1);
    border: 1px solid rgba(var(--primary-color-rgb), 0.3);
    transition: background-color 0.2s ease;
  }
  .calendar-slot:hover {
    background-color: rgba(var(--primary-color-rgb), 0.2);
  }
  .calendar-slot .time {
    font-weight: 700;
    font-size: 1.1rem;
    color: var(--primary-color);
    min-width: 60px; /* Ensure time takes enough space */
  }
  .calendar-slot .patient-info {
    flex-grow: 1;
    margin-left: 1rem;
    font-weight: 500;
  }
  .calendar-slot .btn-consult {
    flex-shrink: 0;
  }


  /* Responsive adjustments */
  @media (max-width: 768px) {
    .card-header h1 {
      font-size: 1.5rem !important;
    }
    .card-header .header-item {
      font-size: 1rem !important;
    }
    .card-header i {
      font-size: 1.5rem !important;
    }
    .gender-btn label {
      flex: 1 1 100%; /* Stack radio buttons on small screens */
    }
    .btn {
      width: 100%; /* Full width buttons on small screens */
      margin-bottom: 0.5rem;
    }
    .d-flex.gap-2 {
      flex-direction: column;
    }
    .dataTables_filter, .dataTables_length {
      text-align: center !important;
    }
    .dataTables_filter input, .dataTables_length select {
      width: 100%;
      margin-bottom: 0.5rem;
    }
    .calendar-slot {
      flex-direction: column;
      align-items: flex-start;
    }
    .calendar-slot .patient-info {
      margin-left: 0;
      margin-top: 0.5rem;
      text-align: center;
      width: 100%;
    }
    .calendar-slot .time {
      width: 100%;
      text-align: center;
    }
    .calendar-slot .btn-consult {
      width: 100%;
      margin-top: 0.5rem;
    }
  }

  /* Améliorations responsives pour SweetAlert2 */
  @media (max-width: 768px) {
    .swal2-popup {
      width: 90vw !important; /* Occupe 90% de la largeur de l'écran */
      font-size: 0.9rem;   /* Réduit la taille de la police globale */
    }
    .swal2-title {
      font-size: 1.25rem !important; /* Réduit la taille du titre */
    }
    .swal2-html-container {
      font-size: 1rem !important; /* Ajuste la taille du texte principal */
    }
    .swal2-actions {
      flex-wrap: wrap; /* Permet aux boutons de passer à la ligne */
      gap: 0.5rem;
    }
    .swal2-confirm, .swal2-cancel {
      width: 100% !important; /* Les boutons prennent toute la largeur */
      margin: 0.25rem 0 !important; /* Empile les boutons verticalement */
    }
  }
</style>
</head>
<body>

<nav class="navbar navbar-dark fixed-top">
  <div class="container-fluid d-flex align-items-center">
    <button class="navbar-toggler" type="button" data-bs-toggle="offcanvas" data-bs-target="#settingsOffcanvas">
      <i class="fas fa-bars"></i>
    </button>
    <a class="navbar-brand ms-auto d-flex align-items-center"
       href="{{ url_for('accueil.accueil') }}">
      <i class="fas fa-home me-2"></i> {# Home Icon (original color) #}
      <i class="fas fa-heartbeat me-2"></i>EasyMedicaLink {# Heartbeat Icon (original color) #}
    </a>
  </div>
</nav>

<div class="offcanvas offcanvas-start" tabindex="-1" id="settingsOffcanvas">
  <div class="offcanvas-header text-white">
    <h5 class="offcanvas-title"><i class="fas fa-cog me-2"></i>Paramètres</h5>
    <button type="button" class="btn-close btn-close-white" data-bs-dismiss="offcanvas"></button>
  </div>
  <div class="offcanvas-body">
    <div class="d-flex gap-2 mb-4">
      <a href="{{ url_for('login.logout') }}" class="btn btn-outline-secondary flex-fill">
        <i class="fas fa-sign-out-alt me-2" style="color: #DC143C;"></i>Déconnexion {# Crimson Sign-out Icon #}
      </a>
    </div>
  </div>
</div>

<div class="container-fluid my-4">
  <div class="row justify-content-center">
    <div class="col-12">
      <div class="card shadow-lg">
        <div class="card-header py-3 text-center">
          <h1 class="mb-2 header-item">
            <i class="fas fa-hospital me-2"></i> {# Hospital Icon (original color) #}
            {{ config.nom_clinique or config.cabinet or 'NOM CLINIQUE/CABINET/CENTRE MEDICAL' }}
          </h1>
          <div class="d-flex justify-content-center gap-4 flex-wrap">
            <div class="d-flex align-items-center header-item">
              <i class="fas fa-user me-2"></i><span>{{ logged_in_doctor_name if logged_in_doctor_name and logged_in_doctor_name != 'None' else config.doctor_name or 'NOM MEDECIN' }}</span> {# Simple User Icon #}
            </div>
            <div class="d-flex align-items-center header-item">
              <i class="fas fa-map-marker-alt me-2"></i><span>{{ config.location or 'LIEU' }}</span> {# Map Marker Icon (original color) #}
            </div>
          </div>
          <p class="mt-2 header-item">
            <i class="fas fa-calendar-day me-2"></i>{{ today }} {# Calendar Icon (original color) #}
          </p>
          <p class="mt-2 header-item">
            <i class="fas fa-calendar-check me-2" style="color: #FFFFFF;"></i>Prise de RDV et Salle d’Attente {# Green Calendar Check Icon #}
          </p>
        </div>
      </div>
    </div>
  </div>
</div>

<div class="container-fluid my-4">
  <div class="row justify-content-center">
    <div class="col-12 col-lg-10">
      <button id="togglePatientForm" class="btn btn-secondary btn-sm mb-3 d-block mx-auto">
        <i class="fas fa-eye-slash me-1"></i>Masquer le formulaire
      </button>
      <div id="patientFormContainer">
        <div class="d-flex align-items-center mb-3">
          <div class="step-circle active" id="step1Circle">1</div>
          <div class="step-line" id="line1"></div>
          <div class="step-circle" id="step2Circle">2</div>
        </div>
        <form method="POST" class="card p-4 shadow-sm" id="wizardForm">
          <div id="step1">
            <h4 class="text-primary mb-3"><i class="fas fa-user me-2" style="color: #007BFF;"></i>Identité Patient</h4> {# Blue User Icon #}
            <div class="row g-3">
              <div class="col-md-4 floating-label">
                <input list="idlist" name="patient_id" id="patient_id"
                       value="{{ edit_row['ID'] if edit_row is defined else '' }}"
                       class="form-control" placeholder=" " required>
                <label for="patient_id"><i class="fas fa-id-card me-2" style="color: #6C757D;"></i>ID Patient</label> {# Gray ID Card Icon #}
                <datalist id="idlist">
                  {% for pid in patients %}
                  <option value="{{ pid }}">
                  {% endfor %}
                </datalist>
              </div>
              <div class="col-md-4 floating-label">
                <input type="text" name="patient_nom" id="patient_nom"
                      value="{{ edit_row['Nom'] if edit_row is defined else '' }}"
                      class="form-control" placeholder=" " required>
                <label for="patient_nom"><i class="fas fa-user me-2" style="color: #007BFF;"></i>Nom</label> {# Blue User Icon #}
              </div>
              <div class="col-md-4 floating-label">
                <input type="text" name="patient_prenom" id="patient_prenom"
                      value="{{ edit_row['Prenom'] if edit_row is defined else '' }}"
                      class="form-control" placeholder=" " required>
                <label for="patient_prenom"><i class="fas fa-user-tag me-2" style="color: #17A2B8;"></i>Prénom</label> {# Cyan User Tag Icon #}
              </div>
              <div class="col-12">
                <div class="gender-btn">
                  <input type="radio" id="genderM" name="patient_gender" value="Masculin" required
                         {% if edit_index is defined and edit_row['Sexe']=='Masculin' %}checked{% endif %}>
                  <label for="genderM"><i class="fas fa-mars" style="color: #007BFF;"></i> Masculin</label> {# Blue Mars Icon #}
                  <input type="radio" id="genderF" name="patient_gender" value="Féminin"
                         {% if edit_index is defined and edit_row['Sexe']=='Féminin' %}checked{% endif %}>
                  <label for="genderF"><i class="fas fa-venus" style="color: #FF69B4;"></i> Féminin</label> {# Pink Venus Icon #}
                  <input type="radio" id="genderO" name="patient_gender" value="Autre"
                         {% if edit_index is defined and edit_row['Sexe'] not in ['Masculin','Féminin'] %}checked{% endif %}>
                  <label for="genderO"><i class="fas fa-genderless" style="color: #6A5ACD;"></i> Autre</label> {# Slate Blue Genderless Icon #}
                </div>
              </div>
              <div class="col-md-4 floating-label">
                <input type="date" name="patient_dob" id="patient_dob"
                       value="{{ edit_row['DateNaissance'] if edit_row is defined else '' }}"
                       class="form-control" placeholder=" " required>
                <label for="patient_dob"><i class="fas fa-calendar-alt me-2" style="color: #FFB6C1;"></i>Date de naissance</label> {# Light Pink Calendar Icon #}
              </div>
              <div class="col-md-4 floating-label">
                <input type="tel" name="patient_phone" id="patient_phone"
                       value="{{ edit_row['Téléphone'] if edit_row is defined else '' }}"
                       class="form-control" placeholder=" " required>
                <label for="patient_phone"><i class="fas fa-phone me-2" style="color: #28A745;"></i>Téléphone</label> {# Green Phone Icon #}
              </div>
              <div class="col-md-4 floating-label">
                <input type="text" name="patient_ant" id="patient_ant"
                       value="{{ edit_row['Antécédents'] if edit_row is defined else '' }}"
                       class="form-control" placeholder=" ">
                <label for="patient_ant"><i class="fas fa-file-medical me-2" style="color: #DA70D6;"></i>Antécédents médicaux</label> {# Orchid File Medical Icon #}
              </div>
            </div>
            <button type="button" id="toStep2" class="btn btn-primary mt-3 w-100"> Suivant <i class="fas fa-arrow-right ms-1"></i>
            </button>
          </div>
          <div id="step2" class="d-none">
            <h4 class="text-primary mb-3"><i class="fas fa-calendar-plus me-2" style="color: #4CAF50;"></i>Détails du RDV</h4> {# Green Calendar Plus Icon #}
            <div class="row g-3">
              <div class="col-md-6 floating-label">
                <input type="date" name="rdv_date" id="rdv_date"
                       value="{{ edit_row['Date'] if edit_row is defined else (filt_date or iso_today) }}"
                       class="form-control" placeholder=" " required>
                <label for="rdv_date"><i class="fas fa-calendar-day me-2" style="color: #FFC107;"></i>Date</label> {# Amber Calendar Day Icon #}
              </div>
              <div class="col-md-6 floating-label">
                <select name="medecin_select" id="medecin_select" class="form-select" required>
                  <option value="">Sélectionnez un médecin</option>
                  {% for doctor in doctors %}
                  <option value="{{ doctor.email }}"
                    {% if edit_row is defined and doctor.email == edit_row['Medecin_Email'] %}selected{% endif %}>
                    {{ doctor.prenom }} {{ doctor.nom }}
                  </option>
                  {% endfor %}
                </select>
                <label for="medecin_select"><i class="fas fa-user-md me-2" style="color: #ADD8E6;"></i>Médecin</label>
              </div>
              <div class="col-12 floating-label"> {# Changed to col-12 as date/doctor take full row #}
                <select name="rdv_time" id="rdv_time" class="form-select" required>
                  <option value="">Sélectionnez une heure</option>
                  {# Options will be dynamically loaded by JavaScript based on `all_possible_slots` #}
                </select>
                <label for="rdv_time"><i class="fas fa-clock me-2" style="color: #20B2AA;"></i>Heure</label> {# Light Sea Green Clock Icon #}
              </div>
            </div>
            <div class="d-flex justify-content-between mt-3"> <button type="button" id="backStep1" class="btn btn-outline-secondary">
                <i class="fas fa-arrow-left me-1"></i>Précédent
              </button>
              <button class="btn btn-success">
                <i class="fas fa-check-circle me-1"></i>Confirmer le RDV
              </button>
            </div>
          </div>
        </form>
      </div>
    </div>
  </div>

  <div class="row mt-4 justify-content-center">
    <div class="col-12 col-md-8 d-flex justify-content-between align-items-center mb-2 flex-wrap gap-2">
      <form method="get" class="d-flex gap-1 flex-grow-1">
        <input type="date" name="date" value="{{ filt_date }}" class="form-control form-control-sm">
        <button class="btn btn-sm btn-primary"><i class="fas fa-filter me-1" style="color: #FFFFFF;"></i>Filtrer</button> {# White Filter Icon #}
      </form>
      <div class="d-flex gap-2">
        <a href="{{ url_for('rdv.rdv_home') }}" class="btn btn-outline-secondary btn-sm"><i class="fas fa-list me-1" style="color: #6C757D;"></i>Tous</a> {# Gray List Icon #}
        <a href="{{ url_for('rdv.rdv_home', date=iso_today) }}" class="btn btn-secondary btn-sm">
          <i class="fas fa-calendar-day me-1"></i>RDV du Jour
        </a>
      </div>
    </div>
  </div>

  {# NEW SECTION: Today's Appointments Calendar Card #}
  <div class="row justify-content-center mt-4">
    <div class="col-12 col-lg-10">
      <div class="card p-3 shadow-sm calendar-card">
<h4 class="text-primary m-0 mb-3">
          <i class="fas fa-calendar-alt me-2" style="color: #FFB6C1;"></i>Rendez-vous du Jour ({{ today }}) {# Light Pink Calendar Icon #}
        </h4>
        {% if today_rdv %}
          {% for r in today_rdv %}
            <div class="calendar-slot">
              <span class="time">{{ r["Heure"] }}</span>
              <span class="patient-info">{{ r["Nom"] }} {{ r["Prenom"] }} - {{ r["Medecin_Email"] }}</span>
              <a href="{{ url_for('rdv.consult_rdv', index=r.original_df_index) }}" {# C'EST LE CHANGEMENT CLÉ #}
                 class="btn btn-sm btn-success btn-consult" title="Passer à la consultation" data-bs-toggle="tooltip">
                <i class="fas fa-stethoscope" style="color: #FFFFFF;"></i>Consultation
              </a>
            </div>
          {% endfor %}
        {% else %}
          <p class="text-center text-muted">Aucun rendez-vous pour aujourd'hui.</p>
        {% endif %}
      </div>
    </div>
  </div>
  {# END NEW SECTION #}

  <div class="row justify-content-center">
    <div class="col-12">
      <div class="card p-3 shadow-sm">
        <h4 class="text-primary m-0 mb-3"><i class="fas fa-calendar-check me-2" style="color: #4CAF50;"></i>Rendez-vous Confirmés</h4> {# Green Calendar Check Icon #}
        <div class="table-responsive">
          <table id="rdvTable" class="table table-striped table-bordered align-middle">
            <thead>
              <tr>
                <th>#</th><th>ID</th><th>Nom</th><th>Prénom</th><th>Sexe</th><th>Âge</th><th>Téléphone</th><th>Antéc.</th><th>Date</th><th>Heure</th><th>Médecin</th><th>Ordre</th><th>Actions</th>
              </tr>
            </thead>
            <tbody>
              {# IMPORTANT: Ici, la boucle utilise 'r' et non 'idx, r' car 'df' est maintenant une liste de dictionnaires avec 'original_df_index' #}
              {% for r in df %}
              <tr>
                {# Utilisez loop.index0 pour le numéro de ligne affiché si vous voulez un index séquentiel #}
                <td>{{ loop.index0 + 1 }}</td><td>{{ r["ID"] }}</td><td>{{ r["Nom"] }}</td><td>{{ r["Prenom"] }}</td><td>{{ r["Sexe"] }}</td>
                <td>{{ r["Âge"] }}</td><td>{{ r["Téléphone"] }}</td><td>{{ r["Antécédents"] }}</td>
                <td>{{ r["Date"] }}</td><td>{{ r["Heure"] }}</td><td>{{ r["Medecin_Email"] }}</td>
                <td>{{ r["Num Ordre"] }}</td><td class="text-center">
                  <a href="{{ url_for('rdv.consult_rdv', index=r.original_df_index) }}"
                     class="btn btn-sm btn-success me-1" title="Consultation" data-bs-toggle="tooltip">
                    <i class="fas fa-stethoscope" style="color: #FFFFFF;"></i>
                  </a>
                  <a href="{{ url_for('rdv.edit_rdv', index=r.original_df_index) }}"
                     class="btn btn-sm btn-warning me-1" title="Modifier" data-bs-toggle="tooltip">
                    <i class="fas fa-pen" style="color: #FFFFFF;"></i>
                  </a>
                  <button class="btn btn-sm btn-danger me-1 delete-rdv-btn"
                          data-index="{{ r.original_df_index }}"
                          title="Supprimer" data-bs-toggle="tooltip">
                    <i class="fas fa-trash" style="color: #FFFFFF;"></i> {# MODIFIÉ: Icône de poubelle blanche #}
                  </button>
                  {# WhatsApp Button #}
                  <a href="#" class="btn btn-sm btn-info whatsapp-btn"
                     data-phone="{{ r['Téléphone'] }}"
                     data-patient-name="{{ r['Nom'] }} {{ r['Prenom'] }}"
                     data-rdv-date="{{ r['Date'] }}"
                     data-rdv-time="{{ r['Heure'] }}"
                     title="Envoyer rappel WhatsApp" data-bs-toggle="tooltip">
                    <i class="fab fa-whatsapp" style="color: #FFFFFF;"></i>
                  </a>
                </td>
              </tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
</div>

<footer class="text-center py-3 small">
  SASTOUKA DIGITAL © 2025 • <i class="fas fa-envelope me-1" style="color: #ADD8E6;"></i>sastoukadigital@gmail.com • <i class="fab fa-whatsapp me-1" style="color: #25D366;"></i>+212652084735 {# Light Blue Envelope and Green WhatsApp Icon #}
</footer>

<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
<script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
<script src="https://cdn.datatables.net/1.13.7/js/dataTables.bootstrap5.min.js"></script>

<script>
  $(function(){ $('#rdvTable').DataTable({order:[[10,'asc']]}); }); /* Note: Order by column 10 (Médecin) after adding it */
</script>

<script>
// Function to show a custom confirmation dialog using SweetAlert2
function showConfirmDialog(message, confirmUrl) {
  Swal.fire({
    title: 'Êtes-vous sûr ?',
    text: message,
    icon: 'warning',
    showCancelButton: true,
    confirmButtonColor: '#3085d6',
    cancelButtonColor: '#d33',
    confirmButtonText: 'Oui, supprimer !',
    cancelButtonText: 'Annuler'
  }).then((result) => {
    if (result.isConfirmed) {
      window.location.href = confirmUrl;
    }
  });
}

document.addEventListener('DOMContentLoaded',()=>{
  const toggleBtn=document.getElementById('togglePatientForm');
  const formCont=document.getElementById('patientFormContainer');
  toggleBtn.onclick=()=>{
    formCont.classList.toggle('d-none');
    toggleBtn.innerHTML=formCont.classList.contains('d-none')?
      '<i class="fas fa-eye me-1"></i>Afficher le formulaire':
      '<i class="fas fa-eye-slash me-1"></i>Masquer le formulaire';
  };

  /* Wizard navigation */
  const step1=document.getElementById('step1'),step2=document.getElementById('step2');
  const s1c=document.getElementById('step1Circle'),s2c=document.getElementById('step2Circle'),line1=document.getElementById('line1');
  document.getElementById('toStep2').onclick=()=>{
    // Basic validation for step 1 before proceeding
    const patientId = document.getElementById('patient_id').value;
    const patientNom = document.getElementById('patient_nom').value;
    const patientPrenom = document.getElementById('patient_prenom').value;
    const patientGenderM = document.getElementById('genderM').checked;
    const patientGenderF = document.getElementById('genderF').checked;
    const patientGenderO = document.getElementById('genderO').checked;
    const patientDob = document.getElementById('patient_dob').value;
    const patientPhone = document.getElementById('patient_phone').value;
    const patientAnt = document.getElementById('patient_ant').value;

    if (!patientId || !patientNom || !patientPrenom || !(patientGenderM || patientGenderF || patientGenderO) || !patientDob || !patientPhone || !patientAnt) {
      Swal.fire({
        icon: 'error',
        title: 'Champs manquants',
        text: 'Veuillez remplir tous les champs de l\'identité du patient avant de continuer.',
        confirmButtonText: 'OK'
      });
      return; // Stop function if validation fails
    }

    step1.classList.add('d-none');step2.classList.remove('d-none');
    s1c.classList.remove('active');line1.classList.add('active');s2c.classList.add('active');
  };
  document.getElementById('backStep1').onclick=()=>{
    step2.classList.add('d-none');step1.classList.remove('d-none');
    s2c.classList.remove('active');line1.classList.remove('active');s1c.classList.add('active');
  };

  /* Tooltips */
  new bootstrap.Tooltip(document.body,{selector:'[data-bs-toggle="tooltip"]'});

  /* WhatsApp button functionality */
  document.querySelectorAll('.whatsapp-btn').forEach(button => {
    button.addEventListener('click', function(e) {
      e.preventDefault();
      const phone = this.dataset.phone;
      const patientName = this.dataset.patientName;
      const rdvDate = this.dataset.rdvDate;
      const rdvTime = this.dataset.rdvTime;

      // Retrieve clinic/cabinet name, doctor name, and location from the config object
      // This assumes 'config' is available in the global scope or passed to the template
      // Fallback values are provided in case config is not fully loaded or properties are missing
      const clinicName = "{{ config.nom_clinique or config.cabinet or 'Votre Clinique/Cabinet' }}";
      const doctorName = "{{ config.doctor_name or 'Votre Médecin' }}";
      const location = "{{ config.location or 'Votre Adresse' }}";

      const message = `Bonjour ${patientName},\n\nCeci est un rappel de votre rendez-vous au ${clinicName} avec le Dr. ${doctorName} à ${location}.\n\nDate: ${rdvDate}\nHeure: ${rdvTime}\n\nNous vous attendons avec impatience !`;

      // Format phone number for WhatsApp (remove leading + if present, ensure only digits)
      const formattedPhone = phone.replace(/[^0-9]/g, ''); // Remove non-digits
      const whatsappLink = `https://wa.me/${formattedPhone}?text=${encodeURIComponent(message)}`;
      
      window.open(whatsappLink, '_blank');
    });
  });

  // Function to update time slots based on the selected date AND selected doctor
  function updateTimeSlots(selectedDate, selectedDoctorEmail) {
      if (!selectedDoctorEmail) { // If no doctor selected, clear time slots
          const rdvTimeSelect = document.getElementById('rdv_time');
          rdvTimeSelect.innerHTML = '<option value="">Sélectionnez un médecin d\'abord</option>';
          rdvTimeSelect.disabled = true;
          return;
      }
      
      fetch(`/rdv/get_reserved_slots?date=${encodeURIComponent(selectedDate)}&medecin_email=${encodeURIComponent(selectedDoctorEmail)}`)
          .then(response => response.json())
          .then(data => {
              const rdvTimeSelect = document.getElementById('rdv_time');
              rdvTimeSelect.disabled = false; // Enable the select once a doctor is chosen
              const allTimeSlots = data.all_possible_slots; // Utiliser les créneaux possibles renvoyés par le backend

              rdvTimeSelect.innerHTML = ''; // Clear current options
              let hasSelectedOption = false;

              allTimeSlots.forEach(time => {
                  const option = document.createElement('option');
                  option.value = time;
                  option.textContent = time;

                  // Disable if the slot is reserved for the selected date AND doctor
                  if (data.reserved_slots.includes(time)) {
                      option.disabled = true;
                  }

                  // If in edit mode, make sure the currently edited slot is selectable
                  {% if edit_row is defined %}
                  const editRdvTime = "{{ edit_row['Heure'] }}";
                  const editRdvDate = "{{ edit_row['Date'] }}";
                  const editMedecinEmail = "{{ edit_row['Medecin_Email'] }}";
                  
                  if (time === editRdvTime && selectedDate === editRdvDate && selectedDoctorEmail === editMedecinEmail) {
                      option.disabled = false; // Re-enable the slot of the appointment being edited
                      option.selected = true;  // Select the slot of the appointment being edited
                      hasSelectedOption = true;
                  }
                  {% endif %}

                  rdvTimeSelect.appendChild(option);
              });

              // Add default 'Sélectionnez une heure' if no option is selected/pre-selected
              if (!hasSelectedOption && rdvTimeSelect.options.length > 0) {
                  const defaultOption = document.createElement('option');
                  defaultOption.value = '';
                  defaultOption.textContent = 'Sélectionnez une heure';
                  defaultOption.selected = true;
                  defaultOption.disabled = true;
                  rdvTimeSelect.prepend(defaultOption);
              }
          })
          .catch(error => {
              console.error('Erreur lors de la récupération des créneaux réservés:', error);
              const rdvTimeSelect = document.getElementById('rdv_time');
              rdvTimeSelect.innerHTML = '<option value="">Erreur de chargement des heures</option>';
              rdvTimeSelect.disabled = true;
          });
  }

  // Event listener for RDV date change
  const rdvDateInput = document.getElementById('rdv_date');
  const medecinSelect = document.getElementById('medecin_select');

  if (rdvDateInput && medecinSelect) {
      // Trigger update when date or doctor changes
      rdvDateInput.addEventListener('change', function() {
          updateTimeSlots(this.value, medecinSelect.value);
      });
      medecinSelect.addEventListener('change', function() {
          updateTimeSlots(rdvDateInput.value, this.value);
      });

      // Initial call on page load
      updateTimeSlots(rdvDateInput.value, medecinSelect.value);
  }


  // AJAX submission for settings form
  document.getElementById('settingsForm').addEventListener('submit', e=>{
    e.preventDefault();
    fetch(e.target.action,{method:e.target.method,body:new FormData(e.target),credentials:'same-origin'})
      .then(r=>{ if(!r.ok) throw new Error('Échec réseau'); return r; })
      .then(()=>Swal.fire({icon:'success',title:'Enregistré',text:'Paramètres sauvegardés.'}).then(()=>location.reload()))
      .catch(err=>Swal.fire({icon:'error',title:'Erreur',text:err.message}));
  });

});
</script>

<script>
/* Auto-fill patient + gender radio */
document.getElementById('patient_id').addEventListener('change',function(){
  fetch(`/rdv/patient_info/${encodeURIComponent(this.value)}`)
    .then(r=>r.json())
    .then(d=>{
      // Use d.nom and d.prenom directly as they are provided by the backend
      document.getElementById('patient_nom').value    = d.nom || '';
      document.getElementById('patient_prenom').value = d.prenom || '';
      document.getElementById('patient_dob').value    = d.date_of_birth || '';
      document.getElementById('patient_phone').value  = d.phone || '';
      document.getElementById('patient_ant').value    = d.antecedents || '';

      /* reset radios */
      ['genderM','genderF','genderO'].forEach(id=>document.getElementById(id).checked=false);
      if(d.gender==='Masculin')        document.getElementById('genderM').checked=true;
      else if(d.gender==='Féminin')    document.getElementById('genderF').checked=true;
      else if(d.gender)                document.getElementById('genderO').checked=true;
    }); 
});
  /* Fonctionnalité du bouton de suppression de RDV */
  document.querySelectorAll('.delete-rdv-btn').forEach(button => {
    button.addEventListener('click', function(e) {
      e.preventDefault();
      const indexToDelete = this.dataset.index;
      
      // Récupérer les informations du patient pour le message de confirmation
      // Assurez-vous que les cellules du tableau sont dans le bon ordre ou utilisez des data-attributs plus spécifiques
      const row = this.closest('tr');
      const patientNom = row.children[2].textContent;    // Colonne Nom
      const patientPrenom = row.children[3].textContent; // Colonne Prénom
      const rdvDate = row.children[8].textContent;       // Colonne Date
      const rdvHeure = row.children[9].textContent;      // Colonne Heure
      
      const message = `Voulez-vous vraiment supprimer le rendez-vous de ${patientNom} ${patientPrenom} le ${rdvDate} à ${rdvHeure} ?`;
      const confirmUrl = `/rdv/delete/${indexToDelete}`; // L'URL pour la route de suppression

      Swal.fire({
        title: 'Êtes-vous sûr ?',
        text: message,
        icon: 'warning',
        showCancelButton: true,
        confirmButtonColor: '#3085d6',
        cancelButtonColor: '#d33',
        confirmButtonText: 'Oui, supprimer !',
        cancelButtonText: 'Annuler'
      }).then((result) => {
        if (result.isConfirmed) {
          // Si l'utilisateur confirme, nous soumettons un formulaire POST à l'URL de suppression.
          // C'est la méthode sécurisée pour les opérations de suppression.
          const form = document.createElement('form');
          form.method = 'POST';
          form.action = confirmUrl;
          document.body.appendChild(form); // Ajoutez le formulaire au corps du document pour pouvoir le soumettre
          form.submit(); // Soumettez le formulaire
        }
      });
    });
  });
</script>
{% include '_floating_assistant.html' %} 
</body>
</html>
"""
